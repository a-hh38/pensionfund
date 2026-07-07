import requests
import pandas as pd
from openpyxl import load_workbook
from pathlib import Path
import time
from datetime import datetime
from openpyxl import load_workbook


# =====================================================
# CONFIGURATION
# =====================================================

EXCEL_FILE = r"C:\Users\ATHARVA\Desktop\Market_Data.xlsx"

GAINERS_SHEET = "TV_Gainers"
LOSERS_SHEET = "TV_Losers"

# TradingView Scanner Endpoint
URL = "https://scanner.tradingview.com/india/scan"

HEADERS = {
    "User-Agent": "Mozilla/5.0",
    "Content-Type": "application/json"
}

# =====================================================
# Function
# =====================================================

def get_market_movers(order="desc"):

    payload = {
        "filter": [
            {
                "left": "type",
                "operation": "equal",
                "right": "stock"
            }
        ],

        "options": {
            "lang": "en"
        },

        "symbols": {
            "query": {
                "types": []
            },
            "tickers": []
        },

        "columns": [
            "name",
            "description",
            "close",
            "change",
            "change_abs",
            "volume",
            "market_cap_basic",
            "sector"
        ],

        "sort": {
            "sortBy": "change",
            "sortOrder": order
        },

        "range": [0, 99]
    }

    r = requests.post(URL, headers=HEADERS, json=payload)
    r.raise_for_status()

    data = r.json()["data"]

    rows = []

    for item in data:

        rows.append({
            "Ticker": item["s"],
            "Company": item["d"][1],
            "Price": item["d"][2],
            "% Change": item["d"][3],
            "Change": item["d"][4],
            "Volume": item["d"][5],
            "Market Cap": item["d"][6],
            "Sector": item["d"][7]
        })

    return pd.DataFrame(rows)

INTERVAL = 60      # seconds

while True:

    try:
        print("=" * 60)
        print("Run Started :", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

        gainers = get_market_movers("desc")
        losers = get_market_movers("asc")

        # Timestamp column
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        gainers.insert(0, "Timestamp", timestamp)
        losers.insert(0, "Timestamp", timestamp)

        try:
            # Workbook exists
            with pd.ExcelWriter(
                EXCEL_FILE,
                engine="openpyxl",
                mode="a",
                if_sheet_exists="overlay"
            ) as writer:

                workbook = writer.book

                # ---------------------------
                # Gainers
                # ---------------------------
                if GAINERS_SHEET in workbook.sheetnames:
                    ws = workbook[GAINERS_SHEET]
                    start_row = ws.max_row
                else:
                    start_row = 0

                gainers.to_excel(
                    writer,
                    sheet_name=GAINERS_SHEET,
                    startrow=start_row,
                    header=(start_row == 0),
                    index=False
                )

                # ---------------------------
                # Losers
                # ---------------------------
                if LOSERS_SHEET in workbook.sheetnames:
                    ws = workbook[LOSERS_SHEET]
                    start_row = ws.max_row
                else:
                    start_row = 0

                losers.to_excel(
                    writer,
                    sheet_name=LOSERS_SHEET,
                    startrow=start_row,
                    header=(start_row == 0),
                    index=False
                )

        except FileNotFoundError:

            with pd.ExcelWriter(EXCEL_FILE, engine="openpyxl") as writer:
                gainers.to_excel(writer, sheet_name=GAINERS_SHEET, index=False)
                losers.to_excel(writer, sheet_name=LOSERS_SHEET, index=False)

        print(f"Completed : {timestamp}")

    except Exception as e:
        print("ERROR:", e)

    print(f"Sleeping for {INTERVAL} seconds...\n")
    time.sleep(INTERVAL)