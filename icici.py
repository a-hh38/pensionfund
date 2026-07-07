"""
====================================================================
ICICI Pension Fund Portfolio Downloader & Equity Extractor
====================================================================

Features
--------
1. Scrapes all monthly portfolio disclosures
2. Downloads every workbook
3. Reads every worksheet (scheme)
4. Extracts only Equity Instruments
5. Appends into one master Excel
6. Skips duplicates
7. Deletes downloaded files
8. Fully automatic

Author : ChatGPT
"""

import os
import re
import time
import shutil
import json
import logging
from pathlib import Path
from datetime import datetime
from urllib.parse import urljoin
from dateutil.relativedelta import relativedelta
import requests
import pandas as pd
from bs4 import BeautifulSoup

from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter

import xlrd

# ------------------------------------------------------------
# CONFIGURATION
# ------------------------------------------------------------

BASE_URL = "https://www.icicipension.in/"
DISCLOSURE_URL = (
    "https://www.icicipension.in/public_disclosure/portfolio_details"
)
API_URL = "https://api.icicipension.in/get_check/portfolio/{month}/{year}"

HEADERS = {
    "User-Agent":
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 "
        "(KHTML, like Gecko) "
        "Chrome/137.0 Safari/537.36"
}

DOWNLOAD_DIR = Path("downloads")

DOWNLOAD_DIR.mkdir(exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s"
)

# ------------------------------------------------------------
# MASTER COLUMNS
# ------------------------------------------------------------

MASTER_COLUMNS = [
    "Month",
    "Fund",
    "Scheme",
    "Particulars",
    "ISIN",
    "Industry",
    "Quantity",
    "Market Value",
    "% of Portfolio"
]

# ------------------------------------------------------------
# REQUEST SESSION
# ------------------------------------------------------------

session = requests.Session()
session.headers.update(HEADERS)

# ------------------------------------------------------------
# HELPERS
# ------------------------------------------------------------

def clean_text(value):
    """
    Removes unwanted spaces/newlines.
    """
    if value is None:
        return ""

    value = str(value)

    value = value.replace("\n", " ")
    value = value.replace("\r", " ")

    value = re.sub(r"\s+", " ", value)

    return value.strip()

def excel_date_to_string(value):
    """
    Converts Excel serial/date into MMM-YY
    """

    if value is None:
        return ""

    try:

        if isinstance(value, (int, float)):
            dt = xlrd.xldate.xldate_as_datetime(value, 0)
            return dt.strftime("%b-%y")

        text = str(value).strip()

        for fmt in (
            "%d-%m-%Y",
            "%d/%m/%Y",
            "%d-%b-%Y",
            "%B %Y",
            "%b %Y",
            "%Y-%m-%d",
        ):
            try:
                return datetime.strptime(text, fmt).strftime("%b-%y")
            except:
                pass

        return text

    except:
        return str(value)
    

def safe_float(value):

    if value is None:
        return None

    try:

        value = str(value)

        value = value.replace(",", "")
        value = value.replace("%", "")
        value = value.strip()

        if value == "":
            return None

        return float(value)

    except:
        return None
    
def delete_downloads():

    if not DOWNLOAD_DIR.exists():
        return

    for file in DOWNLOAD_DIR.glob("*"):

        try:
            file.unlink()

        except:
            pass

# ------------------------------------------------------------
# SCRAPE DISCLOSURE PAGE
# ------------------------------------------------------------
def get_disclosure_links():

    logging.info("Loading disclosure page...")

    r = session.get(DISCLOSURE_URL, timeout=60)
    r.raise_for_status()

    soup = BeautifulSoup(r.text, "html.parser")

    script = soup.find("script", id="__NEXT_DATA__")

    if script is None:
        raise Exception("__NEXT_DATA__ not found")

    data = json.loads(script.string)

    # DEBUG
    with open("nextdata.json", "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2)

    print("Saved nextdata.json")

    return []
# ------------------------------------------------------------
# DOWNLOAD WORKBOOK
# ------------------------------------------------------------

def download_workbook(file_info):
    """
    Downloads one workbook.

    Returns
    -------
    Path
    """

    url = file_info["url"]

    filename = os.path.basename(url)

    filename = filename.split("?")[0]

    save_path = DOWNLOAD_DIR / filename

    logging.info("Downloading %s", filename)

    with session.get(url, stream=True, timeout=120) as r:

        r.raise_for_status()

        with open(save_path, "wb") as f:

            for chunk in r.iter_content(1024 * 64):

                if chunk:

                    f.write(chunk)

    return save_path


# ------------------------------------------------------------
# OPEN XLS WORKBOOK
# ------------------------------------------------------------

def open_workbook(path):
    """
    Opens legacy xls workbook.
    """

    return xlrd.open_workbook(path)


# ------------------------------------------------------------
# READ SHEET
# ------------------------------------------------------------

def sheet_as_rows(sheet):
    """
    Converts worksheet into
    list of rows.
    """

    rows = []

    for r in range(sheet.nrows):

        row = []

        for c in range(sheet.ncols):

            row.append(sheet.cell_value(r, c))

        rows.append(row)

    return rows


# ------------------------------------------------------------
# ITERATE SHEETS
# ------------------------------------------------------------

def workbook_sheets(workbook):
    """
    Generator yielding

        sheet_name,
        rows
    """

    for sheet in workbook.sheets():

        yield sheet.name, sheet_as_rows(sheet)


# ------------------------------------------------------------
# FIND REPORT MONTH
# ------------------------------------------------------------

def extract_month(rows):
    """
    Searches first few rows for
    report month.
    """

    for row in rows[:8]:

        for cell in row:

            value = excel_date_to_string(cell)

            if re.match(r"^[A-Za-z]{3}-\d{2}$", value):
                return value

    return ""


# ------------------------------------------------------------
# FIND SCHEME NAME
# ------------------------------------------------------------
def extract_scheme(rows, sheet_name):

    import re

    for row in rows[:15]:

        text = " ".join(clean_text(x) for x in row).strip()

        if "name of the scheme" not in text.lower():
            continue

        # Keep only everything after "Name of the scheme"
        text = re.split(r"(?i)name\s+of\s+the\s+scheme", text, maxsplit=1)[1].strip()

        return text

    return sheet_name
# ------------------------------------------------------------
# DOWNLOAD ALL WORKBOOKS
# ------------------------------------------------------------
def download_all_workbooks(start_date, end_date):

    downloaded = []

    current = start_date.replace(day=1)

    while current <= end_date:

        month = current.strftime("%b").lower()
        year = current.year

        api = API_URL.format(
            month=month,
            year=year
        )

        logging.info(f"Checking {month.upper()} {year}")

        try:

            r = session.get(
                api,
                timeout=30
            )

            if r.status_code != 200:
                current += relativedelta(months=1)
                continue

            obj = r.json()

            if not obj.get("data"):
                current += relativedelta(months=1)
                continue

            logging.info(
                f"Found disclosure for {month.upper()} {year}"
            )

            for item in obj["data"]:

                attrs = item["attributes"]

                scheme = attrs["Scheme_Details"]["data"]

                fileinfo = scheme["attributes"]

                url = fileinfo["url"]

                filename = fileinfo["name"]

                save_path = DOWNLOAD_DIR / filename

                file = session.get(
                    url,
                    timeout=120
                )

                file.raise_for_status()

                with open(save_path, "wb") as f:
                    f.write(file.content)

                downloaded.append(save_path)

        except Exception as e:

            logging.exception(e)

        current += relativedelta(months=1)

    logging.info(
        f"Downloaded {len(downloaded)} workbook(s)."
    )

    return downloaded
# ------------------------------------------------------------
# FIND EQUITY SECTION
# ------------------------------------------------------------

def find_equity_start(rows):

    for i, row in enumerate(rows):

        text = " ".join(clean_text(x).lower() for x in row)

        if "equity instruments" in text:

            # find first row containing ISIN
            for j in range(i + 1, min(i + 15, len(rows))):

                hdr = " ".join(clean_text(x).lower() for x in rows[j])

                if "isin" in hdr:
                    return j + 1

            # fallback
            return i + 1

    return None


# ------------------------------------------------------------
# FIND END OF EQUITY
# ------------------------------------------------------------

def find_equity_end(rows, start):

    stop_words = [
        "subtotal",
        "debt instruments",
        "government securities",
        "mutual fund",
        "alternative investment fund",
        "reit",
        "invit",
        "asset backed",
        "cash",
        "money market",
        "reverse repo",
        "treps",
        "corporate debt",
        "certificate of deposit"

    ]

    for i in range(start, len(rows)):

        text = " ".join(clean_text(x).lower() for x in rows[i])

        if text == "":
            continue

        if "subtotal" in text:
            return i

        for word in stop_words:

            if word in text:
                return i

    return len(rows)


# ------------------------------------------------------------
# VALID EQUITY ROW
# ------------------------------------------------------------

def is_equity_row(row):

    if len(row) < 6:
        return False

    company = clean_text(row[0])

    isin = clean_text(row[1])

    if company == "":
        return False

    if company.lower() == "shares":
        return False

    if company.lower().startswith("subtotal"):
        return False

    if isin == "":
        return False

    return True


# ------------------------------------------------------------
# PARSE EQUITY HOLDINGS
# ------------------------------------------------------------

def parse_equity_rows(
    rows,
    scheme,
    month,
    seen
):

    start = find_equity_start(rows)

    if start is None:
        return []

    end = find_equity_end(rows, start)

    extracted = []

    for row in rows[start:end]:

        if not is_equity_row(row):
            continue

        particulars = clean_text(row[0])
        isin = clean_text(row[1])
        industry = clean_text(row[2])

        qty = safe_float(row[3])
        value = safe_float(row[4])
        weight = safe_float(row[5])

        key = (
            month,
            scheme,
            isin
        )

        if key in seen:
            continue

        seen.add(key)

        extracted.append(
            {
                "Date": month,
                "Fund": "ICICI",
                "Scheme": scheme,
                "Particulars": particulars,
                "ISIN": isin,
                "Industry": industry,
                "Quantity": qty,
                "Market Value": value,
                "% of Portfolio": weight,
            }
        )

    return extracted

# ------------------------------------------------------------
# PARSE ONE SHEET
# ------------------------------------------------------------
def parse_sheet(
    sheet_name,
    rows,
    seen
):

    report_date = extract_month(rows)

    scheme = extract_scheme(
        rows,
        sheet_name
    )

    return parse_equity_rows(
        rows,
        scheme,
        report_date,
        seen
    )

# ------------------------------------------------------------
# PARSE ONE WORKBOOK
# ------------------------------------------------------------

def parse_workbook(path,
                   seen):

    wb = open_workbook(path)

    all_rows = []

    logging.info("Processing %s", path.name)

    for sheet_name, rows in workbook_sheets(wb):

        logging.info("   Sheet : %s", sheet_name)

        holdings = parse_sheet(
            sheet_name,
            rows,
            seen
        )

        logging.info(
            "      %d equity holdings",
            len(holdings)
        )

        all_rows.extend(holdings)

    return all_rows

# ------------------------------------------------------------
# MAIN
# ------------------------------------------------------------
def main(start_date=None, end_date=None):

    if start_date is None:
        start_date = pd.Timestamp(2025, 4, 1)

    if end_date is None:
        end_date = pd.Timestamp.today().replace(day=1)

    print("=" * 70)
    print("ICICI PENSION PORTFOLIO DOWNLOADER")
    print("=" * 70)

    delete_downloads()

    downloaded_files = download_all_workbooks(
        start_date,
        end_date
    )

    if not downloaded_files:

        return pd.DataFrame(
            columns=[
                "Date",
                "Fund",
                "Scheme",
                "Particulars",
                "ISIN",
                "Industry",
                "Quantity",
                "Market Value",
                "% of Portfolio"
            ]
        )

    seen = set()

    master = []

    for file in downloaded_files:

        try:

            rows = parse_workbook(
                file,
                seen
            )

            if rows:

                master.extend(rows)

        except Exception:

            logging.exception(
                "Error processing %s",
                file.name
            )

        finally:

            try:

                os.remove(file)

                logging.info(
                    "Deleted %s",
                    file.name
                )

            except:
                pass

    if not master:

        return pd.DataFrame(
            columns=[
                "Date",
                "Fund",
                "Scheme",
                "Particulars",
                "ISIN",
                "Industry",
                "Quantity",
                "Market Value",
                "% of Portfolio"
            ]
        )

    final_df = pd.DataFrame(master)

    final_df = final_df[
        [
            "Date",
            "Fund",
            "Scheme",
            "Particulars",
            "ISIN",
            "Industry",
            "Quantity",
            "Market Value",
            "% of Portfolio"
        ]
    ]

    final_df = final_df.drop_duplicates(
        subset=[
            "Date",
            "Scheme",
            "ISIN"
        ]
    )

    final_df = final_df.sort_values(
        [
            "Date",
            "Scheme",
            "Particulars"
        ]
    ).reset_index(drop=True)

    print()
    print("=" * 70)
    print(f"Rows : {len(final_df):,}")
    print("=" * 70)

    return final_df

def run(start_date=None, end_date=None):
    return main(start_date, end_date)
# ------------------------------------------------------------
# ENTRY
# ------------------------------------------------------------

if __name__ == "__main__":

    run()