# ==========================================================
# HDFC PENSION PORTFOLIO SCRAPER
# PART 1 - IMPORTS / CONFIGURATION
# ==========================================================

import os
import re
import shutil
from datetime import datetime
from urllib.parse import urljoin

import pandas as pd
from bs4 import BeautifulSoup
from curl_cffi import requests
from openpyxl import Workbook, load_workbook

# ==========================================================
# CONFIGURATION
# ==========================================================

BASE_URL = "https://www.hdfcpension.com"
PAGE_URL = BASE_URL + "/public-disclosures/"

DOWNLOAD_FOLDER = "downloads"
os.makedirs(DOWNLOAD_FOLDER, exist_ok=True)

# ==========================================================
# SESSION
# ==========================================================

session = requests.Session()

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/149.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.8",
    "Referer": PAGE_URL,
    "Upgrade-Insecure-Requests": "1",
}

# ==========================================================
# MASTER FILE COLUMNS
# ==========================================================
MASTER_COLUMNS = [
    "Date",
    "Fund",
    "Scheme",
    "Particulars",
    "ISIN",
    "Industry",
    "Quantity",
    "Market Value",
    "% of Portfolio"
]
# ==========================================================
# CLEAN TEXT
# ==========================================================

def clean(x):

    if x is None:
        return ""

    return re.sub(r"\s+", " ", str(x)).strip()


# ==========================================================
# EXTRACT DATE FROM FILENAME
# ==========================================================

def extract_date(filename):

    filename = filename.lower()

    months = {
        "jan":"01",
        "feb":"02",
        "mar":"03",
        "apr":"04",
        "may":"05",
        "jun":"06",
        "jul":"07",
        "aug":"08",
        "sep":"09",
        "oct":"10",
        "nov":"11",
        "dec":"12"
    }

    m = re.search(r'([a-z]{3,9})[-_ ]?(\d{2})', filename)

    if m:

        mon = m.group(1)[:3]

        if mon in months:

            year = "20" + m.group(2)
            dt = datetime.strptime(f"{year}-{months[mon]}-01", "%Y-%m-%d")
            return dt.strftime("%b-%y")

    return datetime.today().strftime("%b-%y")


# ==========================================================
# DELETE DOWNLOADS
# ==========================================================

def clear_downloads():

    if not os.path.exists(DOWNLOAD_FOLDER):
        return

    for f in os.listdir(DOWNLOAD_FOLDER):

        try:
            os.remove(os.path.join(DOWNLOAD_FOLDER, f))
        except:
            pass


# ==========================================================
# DOWNLOAD ONE FILE
# ==========================================================

def download_file(url):

    filename = url.split("/")[-1]

    path = os.path.join(DOWNLOAD_FOLDER, filename)

    print(f"Downloading {filename}")

    r = session.get(
        url,
        headers=HEADERS,
        impersonate="chrome",
        timeout=60,
        stream=True
    )

    r.raise_for_status()

    with open(path, "wb") as f:
        for chunk in r.iter_content(16384):
            if chunk:
                f.write(chunk)

    return path

# ==========================================================
# PART 2 - FIND HDFC PORTFOLIO DISCLOSURE LINKS
# ==========================================================

def get_disclosure_links(
        start_date,
        end_date
    ):

    print("\nFetching HDFC Public Disclosures page...")

    response = session.get(
        PAGE_URL,
        headers=HEADERS,
        impersonate="chrome",
        timeout=60
    )

    response.raise_for_status()

    soup = BeautifulSoup(response.text, "lxml")

    links = []

    # ------------------------------------------------------
    # Find the Portfolio heading
    # ------------------------------------------------------

    portfolio_heading = None

    for tag in soup.find_all(["h2", "h3", "h4"]):

        text = clean(tag.get_text()).lower()

        if "hdfc pension portfolio" in text:

            portfolio_heading = tag
            break

    if portfolio_heading is None:
        raise Exception("Could not locate HDFC Pension Portfolio section.")

    # ------------------------------------------------------
    # Find the accordion container
    # ------------------------------------------------------

    container = portfolio_heading.find_parent()

    while container is not None:

        if container.find("details") is not None:
            break

        container = container.parent

    if container is None:
        raise Exception("Could not locate portfolio accordion.")

    # ------------------------------------------------------
    # Extract only monthly portfolio links
    # ------------------------------------------------------

    for details in container.find_all("details"):

        title = clean(details.get_text(" ", strip=True)).lower()

        if "financial year" not in title:
            continue

        for a in details.find_all("a", href=True):

            href = a["href"].strip()

            if href.startswith("/"):
                href = urljoin(BASE_URL, href)

            text = clean(a.get_text()).lower()

            if (
                ("monthly update" in text)
                and
                href.lower().endswith((".xls", ".xlsx"))
            ):

                links.append(href)

        # Remove duplicates while preserving order
        # Remove duplicates while preserving order
    links = list(dict.fromkeys(links))

    filtered = []

    for link in links:

        dt = pd.to_datetime(
            extract_date(os.path.basename(link)),
            format="%b-%y"
        )

        if start_date <= dt <= end_date:
            filtered.append(link)

    links = filtered
    print(f"\nFound {len(links)} HDFC monthly workbooks\n")

    for i, link in enumerate(links, start=1):
        print(f"{i:02d}. {os.path.basename(link)}")

    return links


# ==========================================================
# DOWNLOAD ALL MONTHLY FILES
# ==========================================================

def download_all_files(
        start_date,
        end_date
    ):

    links = get_disclosure_links(
       start_date,
        end_date
    )

    downloaded = []

    for link in links:

        try:

            downloaded.append(
                download_file(link)
            )

        except Exception as e:

            print(f"\nFailed : {link}")
            print(e)

    print(f"\nDownloaded {len(downloaded)} workbooks.\n")

    return downloaded


# ==========================================================
# DELETE FILE
# ==========================================================

def delete_file(path):

    try:

        os.remove(path)

        print(f"Deleted {os.path.basename(path)}")

    except Exception as e:

        print(e)


# ==========================================================
# PROCESS ALL DOWNLOADS
# ==========================================================
def process_all_downloads(
        start_date,
        end_date
    ):

    files = download_all_files(
        start_date,
        end_date
    )

    if not files:

        return pd.DataFrame(
            columns=[
                "Date",
                "Fund",
                "Scheme",
                "Particulars",
                "ISIN",
                "Industry",
                "Quantity",
                "Market Value",
                "% of Portfolio"
            ]
        )

    existing_keys = set()

    master = []

    for file in files:

        print("=" * 70)
        print(os.path.basename(file))

        try:

            rows = extract_equity_from_workbook(
                file,
                existing_keys
            )

            master.extend(rows)

        except Exception as e:

            print(e)

        delete_file(file)

    if not master:

        return pd.DataFrame(
            columns=[
                "Date",
                "Fund",
                "Scheme",
                "Particulars",
                "ISIN",
                "Industry",
                "Quantity",
                "Market Value",
                "% of Portfolio"
            ]
        )

    final_df = pd.DataFrame(
        master,
        columns=[
            "Date",
            "Fund",
            "Scheme",
            "Particulars",
            "ISIN",
            "Industry",
            "Quantity",
            "Market Value",
            "% of Portfolio"
        ]
    )

    final_df = final_df.drop_duplicates(
        subset=[
            "Date",
            "Scheme",
            "ISIN"
        ]
    )

    final_df = final_df.sort_values(
        [
            "Date",
            "Scheme",
            "Particulars"
        ]
    ).reset_index(drop=True)

    print(f"\nExtracted {len(final_df):,} rows.")

    return final_df
# ==========================================================
# PART 3 - EXTRACT EQUITY HOLDINGS
# ==========================================================

from openpyxl import load_workbook


def is_section_heading(text):

    text = clean(text).upper()

    headings = [
        "EQUITY INSTRUMENTS",
        "DEBT INSTRUMENTS",
        "MONEY MARKET INSTRUMENTS",
        "MUTUAL FUND",
        "MUTUAL FUNDS",
        "ETF",
        "TREPS",
        "REIT",
        "INVIT",
        "ALTERNATIVE INVESTMENTS",
        "CASH",
        "NET CURRENT ASSETS",
        "TOTAL",
        "GRAND TOTAL",
    ]

    for h in headings:
        if h in text:
            return True

    return False


# ==========================================================
# FIND HEADER ROW
# ==========================================================

def find_header(sheet, start_row):

    for r in range(start_row, min(start_row + 25, sheet.max_row) + 1):

        vals = [
            clean(sheet.cell(r, c).value).lower()
            for c in range(1, sheet.max_column + 1)
        ]

        print(r, vals)

        joined = " ".join(vals)

        if "isin" in joined:
            return r

    return None

# ==========================================================
# MAP COLUMNS
# ==========================================================

def get_column_map(header):

    mapping = {}

    for i, value in enumerate(header):

        val = clean(value).lower()
        if "security" in val or "name" in val:
            mapping["security"] = i

        elif "isin" in val:
            mapping["isin"] = i

        elif "industry" in val:
            mapping["industry"] = i

        elif "quantity" in val:
            mapping["quantity"] = i

        elif "market value" in val or "market" in val:
            mapping["market"] = i

        elif "%" in val:
            mapping["percent"] = i

        elif "rating" in val:
            mapping["rating"] = i

    return mapping


# ==========================================================
# EXTRACT ONE SHEET
# ==========================================================

def extract_sheet(sheet, file_date, existing_keys):

    rows = []

    scheme = clean(sheet.title)

    if "vatsalya" in scheme.lower():
        scheme = "Scheme NPS Vatsalya"

    equity_row = None

    equity_row = None

    for r in range(1, sheet.max_row + 1):

        found = False

        for c in range(1, min(sheet.max_column, 15) + 1):

            value = clean(sheet.cell(r, c).value).upper()

            if "EQUITY INSTRUMENTS" in value:

                equity_row = r
                found = True
                
                break

        if found:
            break

    if equity_row is None:
        return rows
    print(f"\nSheet      : {sheet.title}")
    print(f"Equity Row : {equity_row}")
    # Data starts immediately after EQUITY INSTRUMENTS
    r = equity_row + 1
    while r <= sheet.max_row:

        values = [
            clean(sheet.cell(r, c).value)
            for c in range(1, 8)
        ]

        row_text = " ".join(values).upper()

        if "SUBTOTAL" in row_text:
            break

        # Stop only when the next major asset class begins
        if row_text.startswith("DEBT INSTRUMENTS"):
            break

        if row_text.startswith("MONEY MARKET INSTRUMENTS"):
            break

        if row_text.startswith("CASH/CASH EQUIVALENT"):
            break

        if row_text.startswith("NET CURRENT ASSETS"):
            break

        if values[0] == "":
            r += 1
            continue

        security = values[0]
        isin = values[1]
        industry = values[2]
        quantity = values[3]
        market = values[4]
        percent = values[5]
        rating = values[6]

        key = (file_date, scheme, isin)

        if key not in existing_keys:

            existing_keys.add(key)

            rows.append([
                file_date,          # Date
                "HDFC Pension",     # Fund
                scheme,             # Scheme
                security,           # Particulars
                isin,               # ISIN
                industry,           # Industry
                quantity,           # Quantity
                market,             # Market Value
                percent             # % of Portfolio
            ])

        r += 1

    print(f"{scheme:<35} {len(rows):>5} holdings")
    print(f"{scheme}: returning {len(rows)} rows")

    return rows


# ==========================================================
# PROCESS ONE WORKBOOK
# ==========================================================
def extract_equity_from_workbook(filepath, existing_keys):

    print(f"\nOpening {os.path.basename(filepath)}")

    wb = load_workbook(filepath, data_only=True)

    file_date = extract_date(os.path.basename(filepath))

    all_rows = []

    for sheet in wb.worksheets:

        print(f"Processing -> {sheet.title}")

        try:

            rows = extract_sheet(
                sheet,
                file_date,
                existing_keys
            )

            all_rows.extend(rows)

        except Exception as e:

            print(f"Skipped {sheet.title}")
            print(e)

    print(f"\nExtracted {len(all_rows)} rows.\n")

    return all_rows
# ==========================================================
# MAIN
# ==========================================================
def main(start_date=None, end_date=None):

    if start_date is None:
        start_date = pd.Timestamp(2025, 4, 1)

    if end_date is None:
        end_date = pd.Timestamp.today().replace(day=1)

    print("=" * 75)
    print("HDFC PENSION PORTFOLIO DOWNLOADER")
    print("=" * 75)

    clear_downloads()

    final_df = process_all_downloads(
        start_date,
        end_date
    )

    clear_downloads()

    return final_df

def run(start_date=None, end_date=None):
    return main(start_date, end_date)

# ==========================================================
# ENTRY POINT
# ==========================================================

if __name__ == "__main__":

    try:

        run()

    except KeyboardInterrupt:

        print("\nProcess interrupted by user.")

    except Exception as e:

        print("\nFatal Error")
        print(e)