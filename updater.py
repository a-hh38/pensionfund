import pandas as pd
import re
from absl import run as absl
from axis import run as axis
from dsp import run as dsp
from hdfc import run as hdfc
from icici import run as icici
from kotak import run as kotak
from lic import run as lic
from sbi import run as sbi
from tata import run as tata
from uti import run as uti

OUTPUT_FILE = "Pension_Funds_Master.xlsx"

SCRAPERS = [
    ("ABSL", absl),
    ("Axis", axis),
    ("DSP", dsp),
    ("HDFC", hdfc),
    ("ICICI", icici),
    ("Kotak", kotak),
    ("LIC", lic),
    ("SBI", sbi),
    ("Tata", tata),
    ("UTI", uti),
]
import re

def normalize_scheme(scheme):

    s = str(scheme).strip().upper()

    s = re.sub(r"\s+", " ", s)
    s = s.replace("_", " ")
    s = s.replace("-", " ")

    # OCR fixes
    s = s.replace(" TIER L ", " TIER I ")
    s = s.replace(" TIER LI", " TIER II")

    # --------------------------------------------------
    # Scheme E Tier 1
    # --------------------------------------------------
    if (
        "E TIER I" in s
        or "E TIER 1" in s
        or ("E" in s and "TIER" in s and "II" not in s and "2" not in s)
    ):
        return "Scheme E Tier 1"

    # --------------------------------------------------
    # Scheme E Tier 2
    # --------------------------------------------------
    if (
        "E TIER II" in s
        or "E TIER 2" in s
    ):
        return "Scheme E Tier 2"

    # --------------------------------------------------
    # Tax Saver
    # --------------------------------------------------
    if "TAX" in s and "SAVER" in s:

        if "II" in s or "2" in s:
            return "Scheme Tax Saver Tier 2"

        return "Scheme Tax Saver Tier 1"

    # --------------------------------------------------
    # APY
    # --------------------------------------------------
    if "ATAL" in s:
        return "APY"

    # --------------------------------------------------
    # NPS Lite
    # --------------------------------------------------
    if "LITE" in s:
        return "NPS Lite"

    # --------------------------------------------------
    # NPS Vatsalya
    # --------------------------------------------------
    if "VATSALYA" in s:
        return "NPS Vatsalya"

    # --------------------------------------------------
    # Jeevan Swarna
    # --------------------------------------------------
    if "JEEVAN" in s:
        return "NPS Jeevan Swarna"

    # --------------------------------------------------
    # Akshay Dhara
    # --------------------------------------------------
    if "AKSHAY" in s:
        return "NPS Akshay Dhara"

    # --------------------------------------------------
    # Kuber
    # --------------------------------------------------
    if "KUBER" in s:
        return "Scheme Kuber Equity Tier 1"

    # --------------------------------------------------
    # Dream
    # --------------------------------------------------
    if "DREAM" in s:
        return "Scheme DREAM Tier 1"

    # --------------------------------------------------
    # Swasthya
    # --------------------------------------------------
    if "SWASTHYA" in s:
        return "Scheme SWASTHYA Tier 1"

    # --------------------------------------------------
    # MFMF
    # --------------------------------------------------
    if "MFMF" in s:
        return "Scheme MFMF Tier 1"

    # --------------------------------------------------
    # UPS
    # --------------------------------------------------
    if "UPS" in s:
        return "Scheme UPS CG"

    # --------------------------------------------------
    # NPS Tier II Composite
    # --------------------------------------------------
    if "COMPOSITE" in s:
        return "Scheme NPS Tier II Composite"

    # --------------------------------------------------
    # Corporate CG
    # --------------------------------------------------
    if "CORP" in s or "CORPORATE" in s:
        return "Corporate-CG Scheme"

    # --------------------------------------------------
    # Scheme CG
    # --------------------------------------------------
    if (
        "CENTRAL" in s
        or s == "CG"
        or s == "CENTRAL GOVT"
        or s == "CENTRAL GOVT."
    ):
        return "Scheme CG"

    # --------------------------------------------------
    # Scheme SG
    # --------------------------------------------------
    if (
        "STATE" in s
        or s == "SG"
        or "STATE GOVT" in s
    ):
        return "Scheme SG"
    # --------------------------------------------------
    # Secure Retirement Equity
    # --------------------------------------------------
    if "SECURE" in s and "RET" in s:
        return "Scheme Secure Retirement Equity"

    # --------------------------------------------------
    # Secure Fund
    # --------------------------------------------------
    if "SECURE FUND" in s:
        return "Scheme Secure Fund"

    # --------------------------------------------------
    # Scheme E Tier 2
    # --------------------------------------------------
    if re.search(r"\bE\b.*TIER.*(II|2)\b", s):
        return "Scheme E Tier 2"

    # --------------------------------------------------
    # Scheme TTS
    # --------------------------------------------------
    if "TTS" in s:
        return "Scheme Tax Saver Tier 2"

    # --------------------------------------------------
    # OTHER (LIC)
    # --------------------------------------------------
    return scheme

def run_scrapers(start_date, end_date):

    master = []

    required_columns = [
        "Date",
        "Fund",
        "Scheme",
        "Particulars",
        "ISIN",
        "Industry",
        "Quantity",
        "Market Value",
        "% of Portfolio"
    ]

    for name, scraper in SCRAPERS:

        print(f"\n{'='*60}")
        print(f"Running {name}")
        print(f"{'='*60}")

        try:

            df = scraper(start_date, end_date)

            if df is None or df.empty:
                print("No rows returned.")
                continue

            if "Month" in df.columns and "Date" not in df.columns:
                df = df.rename(columns={"Month": "Date"})

            df["Fund"] = name

            missing = [
                c for c in required_columns
                if c not in df.columns
            ]

            if missing:
                raise ValueError(
                    f"{name} missing columns: {missing}"
                )

            df = df[required_columns]

            master.append(df)

            print(f"Extracted {len(df):,} rows.")

        except Exception as e:

            print(f"{name} FAILED")
            print(e)

    if not master:
        return pd.DataFrame(columns=required_columns)

    final = pd.concat(
        master,
        ignore_index=True
    )

    final["Scheme"] = final["Scheme"].apply(
        normalize_scheme
    )

    final = final[
        final["Scheme"]
        .astype(str)
        .str.strip()
        .str.upper()
        != "OTHER"
    ]

    final = final.drop(
        columns=[
            "Unnamed: 6",
            "Unnamed: 7"
        ],
        errors="ignore"
    )

    final["ISIN"] = (
        final["ISIN"]
        .fillna("")
        .astype(str)
        .str.strip()
        .str.upper()
    )

    final = final[
        final["ISIN"].str.match(
            r"^INE[A-Z0-9]{9}$",
            na=False
        )
    ]

    final = final.drop_duplicates(
        subset=[
            "Date",
            "Fund",
            "Scheme",
            "ISIN"
        ]
    )

    final["_date"] = pd.to_datetime(
        final["Date"],
        format="%b-%y",
        errors="coerce"
    )

    final = final.sort_values(
        [
            "_date",
            "Fund",
            "Scheme",
            "Particulars"
        ]
    ).drop(
        columns="_date"
    ).reset_index(drop=True)

    return final

def update_manual(start_date, end_date):

    final = run_scrapers(
        start_date,
        end_date
    )

    if final.empty:

        print("Nothing extracted.")
        return final

    final["Date"] = pd.to_datetime(
        final["Date"],
        format="%b-%y",
        errors="coerce"
    )

    final.to_excel(
        OUTPUT_FILE,
        index=False
    )

    from openpyxl import load_workbook

    wb = load_workbook(OUTPUT_FILE)
    ws = wb.active

    for cell in ws["A"][1:]:
        cell.number_format = "mmm-yy"

    wb.save(OUTPUT_FILE)
    wb.close()

    print(
        f"\nSaved {len(final):,} rows."
    )

    return final

def get_latest_month():

    if not pd.io.common.file_exists(OUTPUT_FILE):
        return pd.Timestamp(2025, 3, 1)

    df = pd.read_excel(OUTPUT_FILE)

    if df.empty:
        return pd.Timestamp(2025, 3, 1)

    dates = pd.to_datetime(
        df["Date"],
        format="%b-%y",
        errors="coerce"
    )

    return dates.max()

def update_latest():

    latest = get_latest_month()

    start_date = latest + pd.DateOffset(months=1)

    end_date = pd.Timestamp.today().replace(day=1)

    if start_date > end_date:

        print("Already up to date.")
        return pd.DataFrame()

    print(
        f"Updating from "
        f"{start_date.strftime('%b-%y')} "
        f"to "
        f"{end_date.strftime('%b-%y')}"
    )

    existing = (
        pd.read_excel(OUTPUT_FILE)
        if pd.io.common.file_exists(OUTPUT_FILE)
        else pd.DataFrame()
    )

    new_rows = run_scrapers(
        start_date,
        end_date
    )

    if new_rows.empty:

        print("No new rows.")
        return new_rows

    final = pd.concat(
        [
            existing,
            new_rows
        ],
        ignore_index=True
    )

    final = final.drop_duplicates(
        subset=[
            "Date",
            "Fund",
            "Scheme",
            "ISIN"
        ]
    )

    final["_date"] = pd.to_datetime(
        final["Date"],
        format="%b-%y",
        errors="coerce"
    )

    final = final.sort_values(
        [
            "_date",
            "Fund",
            "Scheme",
            "Particulars"
        ]
    ).drop(
        columns="_date"
    ).reset_index(drop=True)

    final["Date"] = pd.to_datetime(
        final["Date"],
        format="%b-%y",
        errors="coerce"
    )

    final.to_excel(
        OUTPUT_FILE,
        index=False
    )

    from openpyxl import load_workbook

    wb = load_workbook(OUTPUT_FILE)
    ws = wb.active

    for cell in ws["A"][1:]:
        cell.number_format = "mmm-yy"

    wb.save(OUTPUT_FILE)
    wb.close()

    print(
        f"Added {len(new_rows):,} rows."
    )

    return final

def update_database(
    mode="latest",
    start_date=None,
    end_date=None
):

    if mode.lower() == "latest":

        return update_latest()

    return update_manual(
        start_date,
        end_date
    )

def main():

    mode = input(
        "Mode (latest/manual): "
    ).strip().lower()

    if mode == "manual":

        start = pd.to_datetime(
            input(
                "Start Month (YYYY-MM): "
            ) + "-01"
        )

        end = pd.to_datetime(
            input(
                "End Month (YYYY-MM): "
            ) + "-01"
        )

        if start > end:

            print("Start month must not be after end month.")
            return

        update_database(
            mode="manual",
            start_date=start,
            end_date=end
        )

    else:

        update_database(
            mode="latest"
        )
        
if __name__ == "__main__":
    main()