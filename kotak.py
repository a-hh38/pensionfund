
from datetime import datetime
from playwright.sync_api import sync_playwright
import os
from pathlib import Path
import  re
import pandas as pd

BASE_URL = "https://www.kotakpensionfund.com/"

API_URL = (
    BASE_URL +
    "kpfservices/docuploads/uploadeddocs"
)

MONTHS = {
    "January": 1,
    "February": 2,
    "March": 3,
    "April": 4,
    "May": 5,
    "June": 6,
    "July": 7,
    "August": 8,
    "September": 9,
    "October": 10,
    "November": 11,
    "December": 12
}

DOWNLOAD_DIR = Path("Kotak_Downloads")
DOWNLOAD_DIR.mkdir(exist_ok=True)


def download_files(context, monthly):

    downloaded = []

    for item in monthly:

        url = BASE_URL + item["file_url"]

        filename = os.path.basename(item["file_url"])

        filename = re.sub(
            r'[<>:"/\\|?*]',
            "_",
            filename
        )

        outfile = DOWNLOAD_DIR / filename

        print(f"Downloading {filename}")

        try:

            response = context.request.get(url)

            if response.status != 200:

                print(
                    f"Skipped {filename} ({response.status})"
                )

                continue

            with open(outfile, "wb") as f:
                f.write(response.body())

            item["local_file"] = outfile

            downloaded.append(item)

        except Exception as e:

            print(
                f"Failed {filename}: {e}"
            )

    print()
    print(f"Downloaded {len(downloaded)} files.")

    return downloaded

def fetch_monthly_files(start_date, end_date):

    with sync_playwright() as p:

        browser = p.chromium.launch(
            headless=False,
            args=["--disable-blink-features=AutomationControlled"]
        )

        context = browser.new_context()
        page = context.new_page()

        page.goto(
            BASE_URL,
            wait_until="domcontentloaded",
            timeout=60000
        )

        page.wait_for_timeout(5000)

        response = page.evaluate("""
        async () => {

            const r = await fetch(
                "https://www.kotakpensionfund.com/kpfservices/docuploads/uploadeddocs",
                {
                    method:"POST",
                    headers:{
                        "Content-Type":"application/json",
                        "Accept":"application/json, text/plain, */*"
                    },
                    body:JSON.stringify({
                        section_name:"Portfolio"
                    })
                }
            );

            return await r.json();

        }
        """)

        docs = response["data"]

        monthly = []

        for item in docs:

            if item.get("financial_category") != "Monthly Portfolio":
                continue

            if "risk" in item["title"].lower():
                continue

            if not item["file_url"].lower().endswith((".xlsx", ".xls")):
                continue

            dt = datetime(
                int(item["publish_year"]),
                MONTHS[item["publish_month"]],
                1
            )

            if not (start_date <= dt <= end_date):
                continue

            item["date"] = dt
            monthly.append(item)

        monthly.sort(key=lambda x: x["date"])

        print(f"Found {len(monthly)} monthly portfolios.")

        downloaded = download_files(
            context,
            monthly
        )

        browser.close()

        return downloaded
     
from openpyxl import load_workbook

def load_workbooks(downloaded):

    workbooks = []

    for item in downloaded:

        print(f"Loading {item['local_file'].name}")

        wb = load_workbook(
            item["local_file"],
            data_only=True
        )

        workbooks.append({

            "date": item["date"],
            "file": item["local_file"],
            "workbook": wb

        })

    print(f"\nLoaded {len(workbooks)} workbooks.")

    return workbooks

def find_equity_section(ws):

    for r in range(1, ws.max_row + 1):

        row = " ".join(
            str(c.value).lower()
            for c in ws[r]
            if c.value is not None
        )

        if "equities and related investments" in row:
            return r

    return None

def parse_sheet(ws, workbook_date):

    start = find_equity_section(ws)

    if start is None:
        return []

    return_rows = []

    shares_row = None

    for r in range(start, min(start + 10, ws.max_row + 1)):

        text = str(
            ws.cell(r, 1).value or ""
        ).strip().lower()

        if text == "shares":
            shares_row = r
            break

    if shares_row is None:
        return []

    r = shares_row + 1

    while r <= ws.max_row:

        name = ws.cell(r, 1).value

        if name is None:
            r += 1
            continue

        name = str(name).strip()

        if not name:
            r += 1
            continue

        lower = name.lower()

        if lower == "shares":
            r += 1
            continue

        if "subtotal" in lower:
            break

        if lower.startswith((
            "government",
            "debt",
            "money market",
            "mutual fund",
            "cash",
            "treps",
            "repo",
            "gold",
            "real estate",
            "reit",
            "invit",
            "alternate"
        )):
            break

        isin = ws.cell(r, 2).value

        if not isin:
            r += 1
            continue

        return_rows.append(
            {
                "Date": workbook_date.strftime("%b-%y"),
                "Fund": "Kotak",
                "Scheme": ws.title,
                "Particulars": name,
                "ISIN": isin,
                "Industry": ws.cell(r, 3).value,
                "Quantity": ws.cell(r, 4).value,
                "Market Value": ws.cell(r, 5).value,
                "% of Portfolio": ws.cell(r, 6).value,
            }
        )

        r += 1

    return return_rows

def main(start_date=None, end_date=None):

    if start_date is None:
        start_date = pd.Timestamp(2025, 4, 1)

    if end_date is None:
        end_date = pd.Timestamp.today().replace(day=1)

    downloaded = fetch_monthly_files(
        start_date,
        end_date
    )

    if not downloaded:

        return pd.DataFrame(
            columns=[
                "Date",
                "Fund",
                "Scheme",
                "Particulars",
                "ISIN",
                "Industry",
                "Quantity",
                "Market Value",
                "% of Portfolio"
            ]
        )

    workbooks = load_workbooks(downloaded)

    master = []

    for wb in workbooks:

        for sheet in wb["workbook"].sheetnames:

            master.extend(
                parse_sheet(
                    wb["workbook"][sheet],
                    wb["date"]
                )
            )

        wb["workbook"].close()

        try:
            os.remove(wb["file"])
        except:
            pass

    if not master:

        return pd.DataFrame(
            columns=[
                "Date",
                "Fund",
                "Scheme",
                "Particulars",
                "ISIN",
                "Industry",
                "Quantity",
                "Market Value",
                "% of Portfolio"
            ]
        )

    df = pd.DataFrame(master)

    df["ISIN"] = (
        df["ISIN"]
        .fillna("")
        .astype(str)
        .str.strip()
        .str.upper()
    )

    df = df[
        df["ISIN"].str.match(
            r"^INE[A-Z0-9]{9}$",
            na=False
        )
    ]

    for col in [
        "Quantity",
        "Market Value",
        "% of Portfolio"
    ]:
        df[col] = pd.to_numeric(
            df[col],
            errors="coerce"
        )

    df = df.drop_duplicates(
        subset=[
            "Date",
            "Scheme",
            "ISIN"
        ]
    )

    df = df.sort_values(
        [
            "Date",
            "Scheme",
            "Particulars"
        ]
    ).reset_index(drop=True)

    return df[
        [
            "Date",
            "Fund",
            "Scheme",
            "Particulars",
            "ISIN",
            "Industry",
            "Quantity",
            "Market Value",
            "% of Portfolio"
        ]
    ]

def run(start_date=None, end_date=None):
    return main(start_date, end_date)

if __name__ == "__main__":
    run()