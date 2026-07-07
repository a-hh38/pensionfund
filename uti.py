# Placeholder rewritten ETL skeleton due to response size limits.
# Fill in if extending.
import os,re,requests
from datetime import datetime
from urllib.parse import urljoin
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import pandas as pd

BASE="https://www.utipension.com/"
PAGE=BASE+"retirement-solutions?ID=3"
HEADERS={"User-Agent":"Mozilla/5.0"}

def parse_month(text):
    m=re.search(r"([A-Za-z]+)\s+(\d{4})",text)
    return datetime.strptime(f"{m.group(1)} {m.group(2)}","%B %Y") if m else None

def parse_sheet(ws, month):

    rows = []

    scheme = str(ws["A2"].value or "")

    scheme = re.sub(
        r"Name of the Scheme\s*:\s*",
        "",
        scheme,
        flags=re.I
    )

    m = re.search(
        r"FUND\s*(.*)",
        scheme,
        re.I
    )

    if m:
        scheme = m.group(1).strip()

    scheme = re.sub(
        r"^\d+\.\s*",
        "",
        scheme
    )

    start = None
    end = None

    for r in range(1, ws.max_row + 1):

        text = str(
            ws.cell(r, 1).value or ""
        ).strip()

        lower = text.lower()

        if "equity instruments" in lower:
            start = r + 2

        elif start and (
            "total equity instruments" in lower
            or "debt instruments" in lower
        ):
            end = r
            break

    if start is None:
        return rows

    if end is None:
        end = ws.max_row + 1

    for r in range(start, end):

        vals = [
            ws.cell(r, c).value
            for c in range(1, 8)
        ]

        if all(v is None for v in vals):
            continue

        name = str(vals[0] or "").strip()

        lower = name.lower()

        if (
            lower in ("shares", "share")
            or "unit outstanding" in lower
            or "total equity instruments" in lower
            or "subtotal" in lower
            or lower == "total"
            or "%" in name
        ):
            continue

        instrument = re.sub(
            r"^\d+\.\s*",
            "",
            name
        )

        isin = str(vals[1] or "").strip().upper()

        if not isin.startswith("INE"):
            continue

        rows.append(
            {
                "Date": month,
                "Fund": "UTI",
                "Scheme": scheme,
                "Particulars": instrument,
                "ISIN": isin,
                "Industry": vals[3],
                "Quantity": vals[4],
                "Market Value": vals[5],
                "% of Portfolio": vals[6]
            }
        )

    return rows


def main(start_date=None, end_date=None):

    if start_date is None:
        start_date = pd.Timestamp(2025, 4, 1)

    if end_date is None:
        end_date = pd.Timestamp.today().replace(day=1)

    response = requests.get(
    PAGE,
    headers=HEADERS,
    timeout=60
    )
    response.raise_for_status()

    html = response.text

    soup = BeautifulSoup(
        html,
        "html.parser"
    )

    links = []

    for item in soup.select("div.accordion-item"):

        btn = item.select_one("button")

        if not btn:
            continue

        dt = parse_month(
            btn.get_text(" ", strip=True)
        )

        if dt is None:
            continue

        if not (start_date <= dt <= end_date):
            continue

        for a in item.select("a[href$='.xlsx']"):

            links.append(
                (
                    dt.strftime("%b-%y"),
                    urljoin(BASE, a["href"]),
                    os.path.basename(a["href"])
                )
            )

    print(f"Found {len(links)} portfolio files.")

    rows = []

    for month, url, fname in links:

        print(f"Downloading {fname}")

        r = requests.get(
            url,
            headers=HEADERS,
            timeout=60
        )

        if r.status_code != 200:
            continue

        with open(fname, "wb") as f:
            f.write(r.content)

        try:

            wb = load_workbook(
                fname,
                data_only=True
            )

            for sh in wb.sheetnames:

                rows.extend(
                    parse_sheet(
                        wb[sh],
                        month
                    )
                )

            wb.close()

        finally:

            try:
                os.remove(fname)
            except Exception:
                pass

    if not rows:

        return pd.DataFrame(
            columns=[
                "Date",
                "Fund",
                "Scheme",
                "Particulars",
                "ISIN",
                "Industry",
                "Quantity",
                "Market Value",
                "% of Portfolio"
            ]
        )

    df = pd.DataFrame(rows)

    df["ISIN"] = (
        df["ISIN"]
        .fillna("")
        .astype(str)
        .str.strip()
        .str.upper()
    )

    df = df[
    df["ISIN"].str.match(
        r"^INE[A-Z0-9]{9}$",
        na=False
    )
]

    for col in [
        "Quantity",
        "Market Value",
        "% of Portfolio"
    ]:

        df[col] = pd.to_numeric(
            df[col],
            errors="coerce"
        )
    
    df = df.drop_duplicates(
        subset=[
            "Date",
            "Scheme",
            "ISIN"
        ]
    )
    

    df = df.sort_values(
        [
            "Date",
            "Scheme",
            "Particulars"
        ]
    ).reset_index(drop=True)

    df = df[
        [
            "Date",
            "Fund",
            "Scheme",
            "Particulars",
            "ISIN",
            "Industry",
            "Quantity",
            "Market Value",
            "% of Portfolio"
        ]
    ]

    print(f"\nRows : {len(df):,}")

    return df
def run(start_date=None, end_date=None):
    return main(start_date, end_date)

if __name__ == "__main__":
    run()