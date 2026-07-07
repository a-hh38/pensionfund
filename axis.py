import os
import requests
import pandas as pd
import pandas as pd
from openpyxl import load_workbook
from pathlib import Path
from datetime import datetime, date
from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook

# ============================================================
# CONFIG
# ============================================================

BASE_API = "https://www.axispensionfund.com/wordpress/wp-json/wp/v2/portfolio_reports/"

DOWNLOAD_DIR = Path("axis_downloads")
DOWNLOAD_DIR.mkdir(exist_ok=True)


HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 Chrome/137.0 Safari/537.36"
    )
}

session = requests.Session()
session.headers.update(HEADERS)

def fy_year(dt):
    return dt.year


# ============================================================
# DOWNLOAD URL
# ============================================================

def get_download_url(item):

    acf = item.get("acf", {})

    possible = [
        "excel_report",
        "xlsx_report",
        "xls_report",
        "pdf_report",
        "report_pdf",
        "download_file",
        "file",
    ]

    for key in possible:

        value = acf.get(key)

        if value:
            return value

    return None


# ============================================================
# DOWNLOAD REPORTS
# ============================================================

def download_reports(start_date, end_date):

    downloaded = []

    current = start_date.replace(day=1)

    while current <= end_date:

        month = current.month
        year = current.year

        api = f"{BASE_API}?month={month}&year={year}"

        print(f"\nQuery : {api}")

        try:

            r = session.get(api, timeout=30)

            if r.status_code != 200:
                current += relativedelta(months=1)
                continue

            data = r.json()

            if not data:
                current += relativedelta(months=1)
                continue

            item = data[0]

            download_url = get_download_url(item)

            if not download_url:
                current += relativedelta(months=1)
                continue

            extension = Path(download_url).suffix.lower()

            if extension not in [".xlsx", ".xls"]:
                extension = ".xlsx"

            filename = DOWNLOAD_DIR / f"{current:%Y_%m}{extension}"

            print(f"Downloading {filename.name}")

            file = session.get(download_url, timeout=60)
            file.raise_for_status()

            filename.write_bytes(file.content)

            downloaded.append(filename)

        except Exception as e:

            print(e)

        current += relativedelta(months=1)

    return downloaded

# ============================================================
# EXTRACT EQUITY
# ============================================================

STOP_WORDS = [
    "debt",
    "debt instruments",
    "government securities",
    "government security",
    "corporate debt",
    "money market",
    "mutual fund",
    "mutual funds",
    "units",
    "reit",
    "invit",
    "cash",
    "cash & cash equivalents",
    "cash and cash equivalents",
    "others",
    "others including",
    "net current assets",
    "current assets",
    "current liabilities",
    "derivatives",
    "asset backed",
    "commercial paper",
    "certificate of deposit",
]


def find_equity_start(rows):

    equity_found = False

    for i, row in enumerate(rows):

        values = [
            str(x).strip().lower()
            for x in row
            if x is not None
        ]

        if "equity instruments" in values:
            equity_found = True
            continue

        if equity_found and "shares" in values:
            return i + 1

    return None


def extract_sheet(ws):

    rows = list(ws.iter_rows(values_only=True))

    start = find_equity_start(rows)

    if start is None:
        return pd.DataFrame()

    records = []

    for row in rows[start:]:

        if row[0] is None:
            continue

        first = str(row[0]).strip()

        ignore_keywords = [
            "cash/cash equivalent",
            "cash equivalent",
            "net current assets",
            "out of above",
            "average maturity",
            "modified duration",
            "yield to maturity",
            "net npa",
            "central government securities",
            "state government loans",
            "aaa / equivalent",
            "lower (below investment grade)",
            "bank fd",
            "money market investment",
            "application pending allotment",
            "equity & equity related investment"
        ]

        first_lower = first.lower()

        if any(k in first_lower for k in ignore_keywords):
            continue

        if first.lower() in STOP_WORDS:
            break

        if first == "":
            continue

        isin = row[1] if len(row) > 1 else None
        industry = row[2] if len(row) > 2 else None
        qty = row[3] if len(row) > 3 else None
        value = row[4] if len(row) > 4 else None
        weight = row[5] if len(row) > 5 else None
        rating = row[6] if len(row) > 6 else None

        if (
            isin is None
            and industry is None
            and qty is None
            and value is None
        ):
            continue

        rating = row[6] if len(row) > 6 else None

        rating_str = "" if rating is None else str(rating).strip()

        # Only keep blank/zero ratings
        if rating_str not in ("", "0", "0.0"):
            continue

        records.append({
            "Particulars": first,
            "ISIN": isin,
            "Industry": industry,
            "Quantity": qty,
            "Market Value": value,
            "% of Portfolio": weight,
            "Ratings": rating,
        })

    df = pd.DataFrame(records)

    if len(df):

        df = df.drop_duplicates()

        df = df.reset_index(drop=True)

    return df


# ============================================================
# PROCESS ONE WORKBOOK
# ============================================================

def process_workbook(path):

    print(f"\nProcessing {path.name}")

    output = {"All Schemes": []}

    if path.suffix.lower() == ".xls":

        excel = pd.ExcelFile(path, engine="xlrd")

        for sheet in excel.sheet_names:

            sheet_lower = sheet.lower()

            if "scheme c" in sheet_lower or "scheme g" in sheet_lower:
                print(f"Skipping {sheet}")
                continue

            df_raw = pd.read_excel(
                path,
                sheet_name=sheet,
                header=None,
                engine="xlrd"
            )

            rows = df_raw.values.tolist()

            start = find_equity_start(rows)

            if start is None:
                continue

            records = []

            for row in rows[start:]:

                first = str(row[0]).strip() if pd.notna(row[0]) else ""

                if first.lower() in STOP_WORDS:
                    break

                if first == "":
                    continue

                rating = row[6] if len(row) > 6 else None
                rating_str = "" if rating is None else str(rating).strip()

                if rating_str not in ("", "0", "0.0"):
                    continue

                records.append({
                    "Scheme": sheet,
                    "Particulars": row[0],
                    "ISIN": row[1] if len(row) > 1 else None,
                    "Industry": row[2] if len(row) > 2 else None,
                    "Quantity": row[3] if len(row) > 3 else None,
                    "Market Value": row[4] if len(row) > 4 else None,
                    "% Portfolio": row[5] if len(row) > 5 else None,
                    "Ratings": rating,
                })

            if records:
                output["All Schemes"].append(pd.DataFrame(records))

    else:

        wb = load_workbook(path, data_only=True)

        for sheet in wb.sheetnames:

            sheet_lower = sheet.lower()

            if "scheme c" in sheet_lower or "scheme g" in sheet_lower:
                print(f"Skipping {sheet}")
                continue

            df = extract_sheet(wb[sheet])

            if len(df):

                df.insert(0, "Scheme", sheet)

                output["All Schemes"].append(df)

    if output["All Schemes"]:
        output["All Schemes"] = pd.concat(
            output["All Schemes"],
            ignore_index=True
        )
    else:
        output = {}

    return output

# ============================================================
# CONSOLIDATE ALL MONTHS
# ============================================================
def consolidate_results(results):

    if not results:

        return pd.DataFrame(
            columns=[
                "Date",
                "Fund",
                "Scheme",
                "Particulars",
                "ISIN",
                "Industry",
                "Quantity",
                "Market Value",
                "% of Portfolio"
            ]
        )

    frames = []

    for month, df in results:

        if df.empty:
            continue

        df = df.copy()

        df.insert(0, "Date", month)
        df.insert(1, "Fund", "Axis")

        df = df.rename(
            columns={
                "% Portfolio": "% of Portfolio"
            }
        )

        df = df.drop(
            columns=["Ratings"],
            errors="ignore"
        )

        required = [
            "Date",
            "Fund",
            "Scheme",
            "Particulars",
            "ISIN",
            "Industry",
            "Quantity",
            "Market Value",
            "% of Portfolio"
        ]

        for col in required:

            if col not in df.columns:
                df[col] = None

        frames.append(df[required])

    if not frames:

        return pd.DataFrame(columns=required)

    final_df = pd.concat(
        frames,
        ignore_index=True
    )

    final_df = final_df.drop_duplicates(
        subset=[
            "Date",
            "Scheme",
            "ISIN"
        ]
    )

    final_df = final_df.sort_values(
        [
            "Date",
            "Scheme",
            "Particulars"
        ]
    ).reset_index(drop=True)

    return final_df
# ============================================================
# MAIN
# ============================================================
def main(start_date=None, end_date=None):

    if start_date is None:
        start_date = pd.Timestamp(2025, 4, 1)

    if end_date is None:
        end_date = pd.Timestamp.today().replace(day=1)

    print("=" * 70)
    print("Axis Pension Fund Downloader")
    print("=" * 70)

    DOWNLOAD_DIR.mkdir(exist_ok=True)

    for file in DOWNLOAD_DIR.rglob("*"):

        if file.is_file():

            try:
                file.unlink()
            except:
                pass

    downloaded = download_reports(
        start_date,
        end_date
    )

    if not downloaded:

        return pd.DataFrame(
        columns=[
            "Date",
            "Fund",
            "Scheme",
            "Particulars",
            "ISIN",
            "Industry",
            "Quantity",
            "Market Value",
            "% of Portfolio"
        ]
        )

    results = []

    for workbook in downloaded:

        output = process_workbook(workbook)

        if "All Schemes" in output:

            month = datetime.strptime(
                workbook.stem,
                "%Y_%m"
            ).strftime("%b-%y")

            results.append(
                (
                    month,
                    output["All Schemes"]
                )
            )

        try:
            workbook.unlink()
        except:
            pass

    final_df = consolidate_results(results)

    print(f"\nRows : {len(final_df):,}")

    return final_df

def run(start_date=None, end_date=None):
    return main(start_date, end_date)
# ============================================================
# ENTRY POINT
# ============================================================

if __name__ == "__main__":

    try:

        run()

    except KeyboardInterrupt:

        print("\nCancelled by user.")

    except Exception as e:

        import traceback

        print("\nERROR OCCURRED\n")

        traceback.print_exc()

        input("\nPress ENTER to exit...")