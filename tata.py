import os
import json
import base64
import requests
from pathlib import Path
import re
from datetime import datetime
from Cryptodome.Cipher import AES
from Cryptodome.Util.Padding import pad, unpad
import pandas as pd
from openpyxl import load_workbook

# ============================================================
# CONFIG
# ============================================================

BASE_URL = "https://tatapensionfund.com"

API_BASE = (
    BASE_URL +
    "/pA7xR23mQb9TgH4L/api"
)

DOWNLOAD_DIR = Path("downloads")
DOWNLOAD_DIR.mkdir(exist_ok=True)

KEY = b"1203199320052011"
IV = b"1203199320052011"

HEADERS = {
    "User-Agent":
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64)",

    "Origin":
        BASE_URL,

    "Referer":
        BASE_URL + "/public-disclosure-and-nav",

    "Content-Type":
        "text/plain",

    "Accept":
        "application/json, text/plain, */*"
}

session = requests.Session()
session.headers.update(HEADERS)


# ============================================================
# CRYPTO
# ============================================================

def encrypt(data: dict) -> str:

    text = json.dumps(
        data,
        separators=(",", ":")
    )

    cipher = AES.new(
        KEY,
        AES.MODE_CBC,
        IV
    )

    encrypted = cipher.encrypt(
        pad(
            text.encode(),
            AES.block_size
        )
    )

    return base64.b64encode(
        encrypted
    ).decode()


def decrypt(ciphertext: str):

    ciphertext = ciphertext.strip('"')

    cipher = AES.new(
        KEY,
        AES.MODE_CBC,
        IV
    )

    decrypted = cipher.decrypt(
        base64.b64decode(ciphertext)
    )

    decrypted = unpad(
        decrypted,
        AES.block_size
    )

    return json.loads(
        decrypted.decode()
    )


# ============================================================
# API
# ============================================================

def api_get(endpoint):

    r = session.get(
        API_BASE + endpoint,
        timeout=60
    )

    r.raise_for_status()

    return decrypt(r.text)


def api_post(endpoint, payload):

    encrypted = encrypt(payload)

    r = session.post(
        API_BASE + endpoint,
        data=encrypted,
        timeout=60
    )

    r.raise_for_status()

    return decrypt(r.text)


# ============================================================
# CATEGORY API
# ============================================================

def get_categories():

    print("Loading categories...")

    categories = api_get(
        "/document-categories"
    )

    print(
        f"Found {len(categories)} categories."
    )

    return categories

# ============================================================
# DOCUMENT API
# ============================================================
def get_documents(category):

    category_id = category["id"]

    print("\n" + "=" * 80)
    print("CATEGORY", category)
    print("=" * 80)

    docs = api_post(
        "/documents?sort=id%20asc",
        {
            "document_category_id": category_id
        }
    )

    print("Returned:", len(docs), "documents")
    if category_id == 2:
        with open(
            "category2.json",
            "w",
            encoding="utf-8"
        ) as f:
            json.dump(
                docs,
                f,
                indent=4,
                ensure_ascii=False
            )

        print("Saved Category 2 to category2.json")

    for d in docs:
        print(
            d["id"],
            d.get("name"),
            d.get("file"),
            d.get("document_sub_category_id")
        )

    return docs

# ============================================================
# PRESIGNED URL
# ============================================================

def get_presigned_url(document):

    path = (
        str(document["document_category_id"])
        + "/"
        + document["file"]
    )

    result = api_post(

        "/presigned-url-website",

        {
            "path": path
        }

    )

    return result["url"]


# ============================================================
# DOWNLOAD
# ============================================================

def download_document(document):

    url = get_presigned_url(document)

    filename = document["file"]

    outfile = DOWNLOAD_DIR / filename

    if outfile.exists():

        print("Already exists:", filename)

        return outfile

    print("Downloading:", filename)

    r = session.get(
        url,
        stream=True,
        timeout=120
    )

    r.raise_for_status()

    with open(outfile, "wb") as f:

        for chunk in r.iter_content(1024 * 1024):

            if chunk:

                f.write(chunk)

    print("Saved:", outfile)

    return outfile


# ============================================================
# DOWNLOAD EVERYTHING
# ============================================================
# ============================================================
# CLEAN DOWNLOADS
# ============================================================

def clear_downloads():

    if not DOWNLOAD_DIR.exists():
        return

    print("Deleting old downloaded files...")

    deleted = 0

    for file in DOWNLOAD_DIR.iterdir():

        if file.is_file():

            try:
                file.unlink()
                deleted += 1

            except Exception as e:
                print(f"Couldn't delete {file.name}: {e}")

    print(f"Deleted {deleted} files.\n")
def download_all_documents(start_date, end_date):

    downloaded = []

    categories = get_categories()

    portfolio_category = next(
        (c for c in categories if c["id"] == 2),
        None
    )

    if portfolio_category is None:
        return []

    docs = get_documents(portfolio_category)

    portfolio_docs = []

    for doc in docs:

        if not doc["file"].lower().endswith(".xlsx"):
            continue

        dt = parse_date(doc["file"])

        if dt == datetime.min:
            continue

        if not (start_date <= dt <= end_date):
            continue

        portfolio_docs.append(doc)

    portfolio_docs.sort(
        key=lambda x: parse_date(x["file"])
    )

    for doc in portfolio_docs:

        try:
            downloaded.append(
                download_document(doc)
            )

        except Exception as e:

            print(doc["file"], e)

    return downloaded
# ============================================================
# PARSER
# ============================================================

OUTPUT_COLUMNS = [
    "Date",
    "Fund Name",
    "Scheme Name",
    "Particulars",
    "ISIN",
    "Industry",
    "Quantity",
    "Market Value",
    "% of Portfolio"
]


SECTION_HEADERS = {
    "equity",
    "equity instruments",
    "debt instruments",
    "government securities",
    "corporate debt",
    "money market instruments",
    "mutual fund units",
    "alternative investment funds",
    "asset backed",
    "cash and cash equivalents",
    "others",
    "shares",
    "bonds",
}


def clean_scheme_name(text):

    if text is None:
        return ""

    text = str(text)

    text = text.replace("Name of the Scheme", "")
    text = text.replace(":", "")

    return text.strip()


def clean_date(text):

    if text is None:
        return ""

    return str(text).replace("Portfolio as on", "").strip()

from datetime import datetime
def inspect_sheet(ws):

    print("=" * 100)
    print(ws.title)
    print("=" * 100)

    for r in range(1, 25):

        vals = []

        for c in range(1, min(15, ws.max_column) + 1):

            vals.append(ws.cell(r, c).value)

        print(r, vals)

def parse_workbook(path, workbook_date):

    print("Parsing", path.name)

    wb = load_workbook(path, data_only=True)

    rows = []

    for ws in wb.worksheets:

        scheme = ""

        for r in range(1, 8):

            val = ws.cell(r, 1).value

            if val and "Name of the Scheme" in str(val):

                scheme = (
                    str(val)
                    .split(":", 1)[1]
                    .strip()
                )

                break

        if not scheme:
            continue

        if "PRIVATE LIMITED" in scheme.upper():

            scheme = (
                scheme.upper()
                .split("PRIVATE LIMITED", 1)[1]
                .strip()
                .title()
            )

        scheme_upper = scheme.upper()

        if (
            "SCHEME A" in scheme_upper
            or "SCHEME C" in scheme_upper
            or "SCHEME G" in scheme_upper
        ):
            continue

        header_row = None

        for row in ws.iter_rows():

            values = [
                str(c.value).strip() if c.value else ""
                for c in row
            ]

            joined = " ".join(values).lower()

            if (
                "name of the instrument" in joined
                and "isin" in joined
                and "quantity" in joined
            ):
                header_row = row[0].row
                break

        if header_row is None:
            continue

        headers = [
            str(c.value).strip().lower()
            if c.value else ""
            for c in ws[header_row]
        ]

        def find_col(*names):

            for i, h in enumerate(headers):

                for n in names:

                    if n.lower() in h:
                        return i

            return None

        instrument_col = find_col("name of the instrument")
        isin_col = find_col("isin")
        industry_col = find_col("industry name")
        qty_col = find_col("quantity")
        mv_col = find_col("mkt value", "market value")
        pct_col = find_col("% of portfolio")

        if instrument_col is None or isin_col is None:
            continue

        in_shares = False

        for row in ws.iter_rows(min_row=header_row + 1):

            instrument = row[instrument_col].value

            if instrument is None:
                continue

            instrument = str(instrument).strip()

            lower = instrument.lower()

            if lower == "shares":
                in_shares = True
                continue

            if in_shares and (
                "government securities" in lower
                or "corporate debt" in lower
                or "money market" in lower
                or "mutual fund" in lower
                or "real estate investment trusts" in lower
                or "alternative investment" in lower
                or "cash & cash equivalents" in lower
                or lower == "bonds"
            ):
                break

            if not in_shares:
                continue

            if (
                lower == ""
                or lower.startswith("total")
                or "grand total" in lower
                or "net current assets" in lower
                or lower in SECTION_HEADERS
            ):
                continue

            isin = row[isin_col].value

            if not isin:
                continue

            rows.append({

                "Date": workbook_date.strftime("%b-%y"),

                "Fund": "Tata",

                "Scheme": scheme,

                "Particulars": instrument,

                "ISIN": str(isin).strip().upper(),

                "Industry": row[industry_col].value if industry_col is not None else None,

                "Quantity": row[qty_col].value if qty_col is not None else None,

                "Market Value": row[mv_col].value if mv_col is not None else None,

                "% of Portfolio": row[pct_col].value if pct_col is not None else None

            })

    wb.close()

    return rows

def build_master(downloaded_files):

    all_rows = []

    columns = [
        "Date",
        "Fund",
        "Scheme",
        "Particulars",
        "ISIN",
        "Industry",
        "Quantity",
        "Market Value",
        "% of Portfolio"
    ]

    for file in downloaded_files:

        try:

            workbook_date = parse_date(file.name)

            if workbook_date == datetime.min:
                print(f"Skipping {file.name} (unable to parse date)")
                continue

            rows = parse_workbook(
                file,
                workbook_date
            )

            all_rows.extend(rows)

        except Exception as e:

            print(file.name, e)

    if not all_rows:
        return pd.DataFrame(columns=columns)

    df = pd.DataFrame(all_rows)

    df["ISIN"] = (
        df["ISIN"]
        .fillna("")
        .astype(str)
        .str.strip()
        .str.upper()
    )

    df = df[
        df["ISIN"].str.match(
            r"^INE[A-Z0-9]{9}$",
            na=False
        )
    ]

    df = df.drop_duplicates(
        subset=[
            "Date",
            "Scheme",
            "ISIN"
        ]
    )

    df["_date"] = pd.to_datetime(
        df["Date"],
        format="%b-%y",
        errors="coerce"
    )

    df = df.dropna(
        subset=["_date"]
    )

    df = (
        df.sort_values(
            [
                "_date",
                "Scheme",
                "Particulars"
            ]
        )
        .drop(columns="_date")
        .reset_index(drop=True)
    )

    return df[columns]
# ============================================================
# HELPERS
# ============================================================

MONTHS = {
    "jan": 1,
    "feb": 2,
    "mar": 3,
    "apr": 4,
    "may": 5,
    "jun": 6,
    "jul": 7,
    "aug": 8,
    "sep": 9,
    "oct": 10,
    "nov": 11,
    "dec": 12,
}


def parse_date(filename):

    name = filename.lower()

    # Look for the actual month name first
    m = re.search(
        r"(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)[-_ ]+(\d{4})",
        name
    )

    if m:

        return datetime(
            int(m.group(2)),
            MONTHS[m.group(1)],
            1
        )

    # Fallback: YYYY-MM
    m = re.search(
        r"(\d{4})[-_](\d{2})",
        name
    )

    if m:

        return datetime(
            int(m.group(1)),
            int(m.group(2)),
            1
        )

    return datetime.min

from datetime import datetime

def main(start_date=None, end_date=None):

    if start_date is None:
        start_date = pd.Timestamp(2025, 4, 1)

    if end_date is None:
        end_date = pd.Timestamp.today().replace(day=1)

    clear_downloads()

    downloaded = download_all_documents(
        start_date,
        end_date
    )

    final_df = build_master(downloaded)

    for file in downloaded:

        try:
            file.unlink()
        except:
            pass

    return final_df

def run(start_date=None, end_date=None):
    return main(start_date, end_date)

if __name__ == "__main__":
    run()