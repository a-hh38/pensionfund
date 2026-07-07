import os
import re
import time
import logging
from pathlib import Path
from datetime import datetime
from urllib.parse import urljoin

import requests
import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import load_workbook

from dateutil.relativedelta import relativedelta

# ==========================================================
# CONFIGURATION
# ==========================================================

BASE_URL = "https://pensionfund.adityabirlacapital.com"
PAGE_URL = BASE_URL + "/public-disclosure"

DOWNLOAD_DIR = "downloads"

os.makedirs(DOWNLOAD_DIR, exist_ok=True)

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/138.0.0.0 Safari/537.36"
    )
}

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s"
)

# ==========================================================
# CONSTANTS
# ==========================================================

FUND_NAME = "ABSL"

SKIP_SHEETS = {
    "Port_C1",
    "Port_C1I",
    "Port_G1",
    "Port_G1I",
    "Port_A I",
    "Port_AI",
    "Port_A1",
}
SKIP_SCHEMES = (
    "a tier",
    "c tier",
    "g tier",
)


OUTPUT_COLUMNS = [
    "Date",
    "Fund",
    "Scheme",
    "Particulars",
    "ISIN",
    "Industry",
    "Quantity",
    "Market Value",
    "% of Portfolio",
]

HEADER_MAP = {
    "name of the instrument": "Particulars",
    "isin no.": "ISIN",
    "isin": "ISIN",
    "industry": "Industry",
    "quantity": "Quantity",
    "market value": "Market Value",
    "% of portfolio": "% of Portfolio",
    "percentage of portfolio": "% of Portfolio",
}

MONTH_PATTERN = re.compile(
    r"(Jan(?:uary)?|Feb(?:ruary)?|Mar(?:ch)?|Apr(?:il)?|May|Jun(?:e)?|Jul(?:y)?|Aug(?:ust)?|Sep(?:tember)?|Oct(?:ober)?|Nov(?:ember)?|Dec(?:ember)?)[ _-]?(\d{2,4})",
    re.I,
)

REIT_STOP_HEADERS = {
    "real estate investment trusts",
    "real estate investment trust",
    "infrastructure investment trusts",
    "infrastructure investment trust",
    "reits",
    "invits",
}
# ==========================================================
# DOWNLOAD LINKS
# ==========================================================

def get_download_links(start_date, end_date):

    logging.info("Fetching Public Disclosure page...")

    response = requests.get(
        PAGE_URL,
        headers=HEADERS,
        timeout=60,
    )

    response.raise_for_status()

    soup = BeautifulSoup(response.text, "html.parser")

    workbooks = []

    for a in soup.find_all("a", href=True):

        href = a["href"]

        if not href.lower().endswith(".xlsx"):
            continue

        if href.startswith("/"):
            href = urljoin(BASE_URL, href)

        filename = os.path.basename(href)

        match = MONTH_PATTERN.search(filename)

        if not match:
            continue

        month_name = match.group(1)[:3].title()
        year = match.group(2)

        if len(year) == 4:
            fmt = "%b %Y"
        else:
            fmt = "%b %y"

        dt = datetime.strptime(
            f"{month_name} {year}",
            fmt
        )

        if not (
            start_date
            <= pd.Timestamp(dt)
            <= end_date
        ):
            continue

        workbooks.append(
            {
                "month": dt.strftime("%b-%y"),
                "url": href,
                "filename": filename,
            }
        )

    if not workbooks:

        return pd.DataFrame(
            columns=[
                "month",
                "url",
                "filename"
            ]
        )

    workbooks = (
        pd.DataFrame(workbooks)
        .drop_duplicates(subset=["url"])
        .sort_values(
            "month",
            ascending=False
        )
        .reset_index(drop=True)
    )

    logging.info(
        f"Selected {len(workbooks)} workbooks."
    )

    return workbooks
# ==========================================================
# DOWNLOAD WORKBOOK
# ==========================================================

def download_workbook(url, filename):

    path = os.path.join(
        DOWNLOAD_DIR,
        filename,
    )

    if os.path.exists(path):

        os.remove(path)

    logging.info(
        f"Downloading {filename}"
    )

    response = requests.get(
        url,
        headers=HEADERS,
        timeout=180,
        stream=True,
    )

    response.raise_for_status()

    with open(path, "wb") as f:

        for chunk in response.iter_content(8192):

            if chunk:

                f.write(chunk)

    return path


# ==========================================================
# DELETE WORKBOOK
# ==========================================================

def cleanup_file(path):

    try:

        if os.path.exists(path):

            os.remove(path)

            logging.info(
                f"Deleted {os.path.basename(path)}"
            )

    except Exception as e:

        logging.warning(e)

# ==========================================================
# NORMALIZE TEXT
# ==========================================================

def normalize(text):

    if text is None:
        return ""

    text = str(text)

    text = text.replace("\n", " ")

    text = re.sub(r"\s+", " ", text)

    return text.strip()


# ==========================================================
# FIND HEADER ROW
# ==========================================================

def find_header_row(ws):

    required = {
        "isin",
        "name of the instrument",
        "quantity",
    }

    for row in range(1, min(ws.max_row, 30) + 1):

        values = []

        for col in range(1, ws.max_column + 1):

            values.append(
                normalize(
                    ws.cell(row, col).value
                ).lower()
            )

        joined = " ".join(values)

        score = 0

        for item in required:

            if item in joined:

                score += 1

        if score >= 2:

            logging.info(
                f"Header row found at {row}"
            )

            return row

    raise Exception("Header row not found")


# ==========================================================
# MAP COLUMNS
# ==========================================================

def map_columns(ws, header_row):

    columns = {}

    for col in range(1, ws.max_column + 1):

        value = normalize(
            ws.cell(header_row, col).value
        ).lower()

        if not value:
            continue

        if "name of the instrument" in value:

            columns["particulars"] = col

        elif "isin" in value:

            columns["isin"] = col

        elif "industry" in value:

            columns["industry"] = col

        elif "quantity" in value:

            columns["quantity"] = col

        elif "market value" in value:

            columns["market_value"] = col

        elif (
            "% of portfolio" in value
            or "percentage of portfolio" in value
        ):

            columns["portfolio"] = col

    required = [
        "particulars",
        "isin",
        "industry",
        "quantity",
        "market_value",
        "portfolio",
    ]

    missing = [
        c
        for c in required
        if c not in columns
    ]

    if missing:

        raise Exception(
            f"Missing columns: {missing}"
        )

    logging.info(columns)

    return columns


# ==========================================================
# SCHEME NAME
# ==========================================================

def get_scheme_name(ws):

    return normalize(
        ws["D3"].value
    )


# ==========================================================
# PORTFOLIO DATE
# ==========================================================
def get_portfolio_date(ws, fallback):

    value = ws["D4"].value

    if isinstance(value, datetime):
        return value.strftime("%b-%y")

    value = normalize(value)

    # Try to parse things like:
    # 31-Mar-2026
    # 31 Mar 2026
    # March 31, 2026

    for fmt in (
        "%d-%b-%Y",
        "%d %b %Y",
        "%B %d, %Y",
        "%d-%B-%Y",
        "%d %B %Y",
    ):
        try:
            return datetime.strptime(value, fmt).strftime("%b-%y")
        except:
            pass

    # If all else fails, use the filename month
    return fallback

# ==========================================================
# EXTRACT ONE SHEET
# ==========================================================

def extract_sheet(ws, portfolio_date):

    sheet_name = ws.title

    if sheet_name in SKIP_SHEETS:

        logging.info(f"Skipping {sheet_name}")

        return []

    logging.info(f"Reading {sheet_name}")

    scheme = get_scheme_name(ws)

    scheme_lower = scheme.lower().strip()

    if any(x in scheme_lower for x in SKIP_SCHEMES):
        logging.info(f"Skipping scheme: {scheme}")
        return []

    header_row = find_header_row(ws)

    cols = map_columns(ws, header_row)

    data = []

    inside_reit_section = False

    for row in range(header_row + 1, ws.max_row + 1):

        # --------------------------------------------------
        # Check every cell for REIT / INVIT section headers
        # --------------------------------------------------

        row_text = " ".join(

            normalize(
                ws.cell(row, col).value
            ).lower()

            for col in range(1, ws.max_column + 1)

        )

        if any(
            x in row_text
            for x in REIT_STOP_HEADERS
        ):

            logging.info(
                "Reached REIT / InvIT section."
            )

            break

        # --------------------------------------------------
        # ISIN FILTER
        # --------------------------------------------------

        isin = normalize(
            ws.cell(
                row,
                cols["isin"]
            ).value
        ).upper()

        if not isin.startswith("INE"):

            continue

        particulars = normalize(

            ws.cell(
                row,
                cols["particulars"]
            ).value

        )

        # Skip debt instruments
        if "%" in particulars:
            continue

        industry = normalize(

            ws.cell(
                row,
                cols["industry"]
            ).value

        )

        quantity = ws.cell(
            row,
            cols["quantity"]
        ).value

        market_value = ws.cell(
            row,
            cols["market_value"]
        ).value

        portfolio = ws.cell(
            row,
            cols["portfolio"]
        ).value

        # completely blank row

        if (
            particulars == ""
            and isin == ""
        ):

            continue

        data.append(

            {
                "Date": datetime.strptime(portfolio_date, "%b-%y"),
                "Fund": FUND_NAME,
                "Scheme": scheme,
                "Particulars": particulars,
                "ISIN No.": isin,
                "Industry": industry,
                "Quantity": quantity,
                "Market Value": market_value,
                "% of Portfolio": portfolio,
            }

        )

    logging.info(
        f"Extracted {len(data)} rows."
    )

    return data

# ==========================================================
# PROCESS ONE WORKBOOK
# ==========================================================

def process_workbook(filepath, fallback_date):

    logging.info(f"Processing {os.path.basename(filepath)}")

    wb = load_workbook(
        filepath,
        data_only=True
    )

    workbook_rows = []

    try:

        for sheet in wb.sheetnames:

            ws = wb[sheet]

            portfolio_date = get_portfolio_date(
                ws,
                fallback_date
            )

            rows = extract_sheet(
                ws,
                portfolio_date
            )

            workbook_rows.extend(rows)

    finally:

        wb.close()

    logging.info(
        f"Workbook contributed {len(workbook_rows)} rows."
    )

    return workbook_rows


# ==========================================================
# SAVE OUTPUT
# ==========================================================
def save_output(rows):

    if not rows:
        return pd.DataFrame(columns=OUTPUT_COLUMNS)

    df = pd.DataFrame(rows)

    df = df.rename(
        columns={
            "ISIN No.": "ISIN"
        }
    )

    df = df[
        OUTPUT_COLUMNS
    ]

    df = df.drop_duplicates(
        subset=[
            "Date",
            "Scheme",
            "ISIN"
        ]
    )

    df = df.sort_values(
        [
            "Date",
            "Scheme",
            "Particulars"
        ]
    ).reset_index(drop=True)

    return df
# ==========================================================
# MAIN
# ==========================================================
def main(start_date=None, end_date=None):

    if start_date is None:
        start_date = pd.Timestamp(2025, 4, 1)

    if end_date is None:
        end_date = pd.Timestamp.today().replace(day=1)

    workbooks = get_download_links(
        start_date,
        end_date
    )

    all_rows = []

    for _, item in workbooks.iterrows():

        filepath = None

        try:

            filepath = download_workbook(
                item["url"],
                item["filename"]
            )

            rows = process_workbook(
                filepath,
                item["month"]
            )

            all_rows.extend(rows)

        except Exception as e:

            logging.exception(e)

        finally:

            if filepath:

                cleanup_file(filepath)

    final_df = save_output(all_rows)

    logging.info("=" * 60)
    logging.info(f"Rows : {len(final_df)}")
    logging.info("=" * 60)

    return final_df

def run(start_date=None, end_date=None):
    return main(start_date, end_date)
# ==========================================================
# ENTRY POINT
# ==========================================================

if __name__ == "__main__":

    run()